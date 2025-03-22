import os
import sys
import threading

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction
from PySide6.QtWidgets import QDialog, QMenu, QProgressDialog, QTableWidgetItem

from Interface.import_ui import Ui_Dialog as ExcelTB

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
TABLE_COLLECTION = "import_data"


class ExcelTable(ExcelTB, QDialog):
    ColInformation = []

    def __init__(
        self, parent=None, database=None, Category: str = None, Source: str = None
    ):
        super().__init__(parent)
        self.setupUi(self)
        self.database = database
        self.category = Category
        self.source = Source
        self.FileName = None
        if not self.database is None:
            self.database.create_table(TABLE_COLLECTION)

        self.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tableWidget.customContextMenuRequested.connect(self.show_context_menu)
        self.tableWidget.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.tableWidget.verticalHeader().customContextMenuRequested.connect(
            self.show_row_menu
        )
        self.tableWidget.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.tableWidget.horizontalHeader().customContextMenuRequested.connect(
            self.show_column_menu
        )
        self.AddToDB = threading.Thread(target=self.save_to_mongodb)

        self.btnUse.clicked.connect(self.just_use_it)
        self.btnSaveUse.clicked.connect(self.thread_save_to_mongodb)

    def show_context_menu(self, pos):
        item = self.tableWidget.itemAt(pos)
        if item:  # التأكد من أن هناك عنصرًا في المكان المضغوط عليه
            row = item.row()
            col = item.column()

            # إنشاء القائمة المنسدلة
            menu = QMenu(self)

            # إضافة خيارات للقائمة
            action_name = QAction("ℹ️ This is Name", self)
            action_phone = QAction("ℹ️ This is Phone Number", self)
            action_account = QAction("ℹ️ This is Account", self)
            action_username = QAction("ℹ️ This is Username", self)
            action_notes = QAction("ℹ️ This is Notes", self)
            action_address = QAction("ℹ️ This is Address", self)
            action_country = QAction("ℹ️ This is Country", self)
            action_city = QAction("ℹ️ This is City", self)
            action_edit = QAction("✏️ Edit", self)
            action_delete = QAction("🗑️ Delete", self)

            action_edit.setShortcut("Ctrl+E")  # تعديل عند الضغط على Ctrl+E
            action_delete.setShortcut("Del")  # حذف عند الضغط على زر الحذف

            action_edit.triggered.connect(lambda: self.edit_cell(row, col))
            action_delete.triggered.connect(lambda: self.delete_cell(row, col))
            action_name.triggered.connect(lambda: self.set_info(row, col, "Name"))
            action_phone.triggered.connect(lambda: self.set_info(row, col, "Number"))
            action_account.triggered.connect(lambda: self.set_info(row, col, "Account"))
            action_username.triggered.connect(
                lambda: self.set_info(row, col, "Username")
            )
            action_notes.triggered.connect(lambda: self.set_info(row, col, "Notes"))
            action_address.triggered.connect(lambda: self.set_info(row, col, "Address"))
            action_country.triggered.connect(lambda: self.set_info(row, col, "Country"))
            action_city.triggered.connect(lambda: self.set_info(row, col, "City"))

            menu.addAction(action_name)
            menu.addAction(action_phone)
            menu.addAction(action_account)
            menu.addAction(action_username)
            menu.addAction(action_notes)
            menu.addAction(action_address)
            menu.addAction(action_country)
            menu.addAction(action_city)
            menu.addAction(action_edit)
            menu.addAction(action_delete)

            # عرض القائمة عند موضع الماوس
            menu.exec(self.tableWidget.mapToGlobal(pos))

    def edit_cell(self, row, col):
        item = self.tableWidget.item(row, col)
        if item:
            item.setText("🔄 Write...")
            item.setBackground(Qt.yellow)  # تغيير لون الخلفية للأصفر بعد التعديل

    def delete_cell(self, row, col):
        self.tableWidget.setItem(row, col, QTableWidgetItem(""))
        item = self.tableWidget.item(row, col)
        if item:
            item.setBackground(Qt.red)  # تغيير لون الخلفية للأحمر بعد الحذف

    def show_row_menu(self, pos):
        row = self.tableWidget.rowAt(pos.y())
        if row >= 0:
            menu = QMenu(self)
            action_delete_row = QAction("🗑️ Delete Row", self)
            action_delete_row.triggered.connect(lambda: self.delete_row(row))
            menu.addAction(action_delete_row)
            menu.exec(self.tableWidget.mapToGlobal(pos))

    def show_column_menu(self, pos):
        col = self.tableWidget.columnAt(pos.x())
        if col >= 0:
            menu = QMenu(self)
            action_delete_col = QAction("🗑️ Delete Col", self)
            action_delete_col.triggered.connect(lambda: self.delete_column(col))
            menu.addAction(action_delete_col)
            menu.exec(self.tableWidget.mapToGlobal(pos))

    def delete_row(self, row):
        self.tableWidget.removeRow(row)

    def delete_column(self, col):
        self.tableWidget.removeColumn(col)

    def set_info(self, row: int, col: int, forCol: str):
        forCol = forCol.lower()
        isBreak = False
        for entry in self.ColInformation:
            for key, value in entry.items():
                if key == col and value != forCol:
                    entry[col] = forCol.lower()
                    isBreak = True
                    break
            if isBreak:
                break
        else:
            # If the column does not exist, append a new entry
            data = {col: forCol}
            self.ColInformation.append(data)
        # Update the info text
        self.txtInfo.setText(f"{forCol} is Col Number {col}")

    def just_use_it(self):
        data_list = []
        for row in range(self.tableWidget.rowCount()):
            row_data = {}
            for col in range(self.tableWidget.columnCount()):
                ColumName = self._getColName(col)
                item = self.tableWidget.item(row, col)
                if item and item.text().strip():  # تجاهل القيم الفارغة
                    if not item is None or not item or len(item) > 1:
                        row_data[ColumName] = item.text().strip()
            if row_data:
                data_list.append(row_data)

        if data_list:
            self.setResult(1)
            self.done(1)
            self.parent().received_data = data_list
        else:
            return None

    def _getColName(self, col):
        for entry in self.ColInformation:
            for key, value in entry.items():
                if key == col:
                    return value

    def thread_save_to_mongodb(self):
        self.progress = QProgressDialog(
            "Waiting for Save all your data in Database", "Cancel", 0, 0, self
        )
        self.progress.setWindowTitle("Please Wait")
        self.progress.setModal(True)
        self.progress.setCancelButton(None)
        self.progress.show()
        self.AddToDB.start()

    def save_to_mongodb(self):
        data_list = []
        for row in range(self.tableWidget.rowCount()):
            row_data = {}
            for col in range(self.tableWidget.columnCount()):
                ColumName = self._getColName(col)
                item = self.tableWidget.item(row, col)
                if item and item.text().strip():
                    if not item is None or not item or len(item) > 1:
                        row_data[ColumName] = item.text().strip()
            if row_data:
                row_data.update(
                    {
                        "filename": self.FileName,
                        "category": self.category,
                        "source": self.source,
                    }
                )
                data_list.append(row_data)

        if data_list:
            for rec in data_list:
                self.database.write_to_database(TABLE_COLLECTION, rec)
            self.setResult(1)
            self.done(1)
            self.parent().received_data = data_list
        else:
            return None

    def read_excel_file(self, filename):
        try:
            FileName = filename.split("/")[-1]
            self.FileName = FileName.replace(" ", "_").replace(".", "_")
            self.setWindowTitle(self.FileName)
            wb = load_workbook(filename)
            sheet = wb.active  # تحديد الورقة النشطة

            rows = sheet.max_row
            cols = sheet.max_column

            self.tableWidget.setRowCount(rows)
            self.tableWidget.setColumnCount(cols)

            # تعبئة الجدول بالبيانات
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if not cell_value is None or cell_value != "" or not cell_value:
                        self.tableWidget.setItem(
                            row - 1, col - 1, QTableWidgetItem(str(cell_value))
                        )
            self.show()
            return self.exec()
        except Exception as e:
            print(f"Error: File '{filename}' not found. {e}")
            return False
